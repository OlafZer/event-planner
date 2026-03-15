#!/usr/bin/env python3
"""generate_invitations.py

Erstellt Einladungs-PDFs aus einer Excel-Gästeliste.
Jede Einladung besteht aus zwei Seiten:
  Seite 1 – optionales Cover-Bild (JPG)
  Seite 2 – Einladungstext aus einem konfigurierbaren Template

Synopsis:
    python scripts/generate_invitations.py --image einladung.jpg [OPTIONEN]

Beispiele:
    # Alle Gäste, Standard-Template, Bild auf Seite 1
    python scripts/generate_invitations.py --image flyer.jpg

    # Nur Kategorie 'Familie', eigenes Template, creme Hintergrund
    python scripts/generate_invitations.py \\
        --image flyer.jpg \\
        --template mein_text.txt \\
        --filter-kategorie Familie \\
        --bg-color '#fffef5' --text-color '#2c2c2c'

    # Vorschau ohne Dateien zu erzeugen
    python scripts/generate_invitations.py --image flyer.jpg --dry-run

Platzhalter im Template:
    {name}          Vorname(n)  (Excel-Spalte 'name')
    {nachname}      Nachname    (Excel-Spalte 'nachname')
    {full_name}     name + nachname kombiniert
    {invite_code}   Persönlicher Einladungscode
    {max_persons}   Maximale Personenzahl
    {invite_link}   Vollständiger Einladungslink
    {email}         E-Mail-Adresse (leer wenn nicht vorhanden)
    {telephone}     Telefonnummer  (leer wenn nicht vorhanden)
"""

import argparse
import re
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# Optional dependencies – helpful error messages when missing
# ---------------------------------------------------------------------------
try:
    import openpyxl
except ImportError:
    sys.exit(
        "openpyxl ist nicht installiert.\n"
        "Bitte installieren: pip install openpyxl"
    )

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas as rl_canvas
except ImportError:
    sys.exit(
        "reportlab ist nicht installiert.\n"
        "Bitte installieren: pip install reportlab"
    )

# ---------------------------------------------------------------------------
# Defaults  (all overridable via CLI parameters)
# ---------------------------------------------------------------------------
DEFAULT_EXCEL = "2026_Party_Gästeleiste.xlsx"
DEFAULT_SHEET = "2026_Party_Gästeleiste"
DEFAULT_OUTPUT_DIR = "invitations"
DEFAULT_BASE_URL = "https://www.zerfowski.de/cgi-bin/app.wsgi"
DEFAULT_EVENT_ID = 2
DEFAULT_FONT_SIZE = 11          # pt
DEFAULT_LINE_SPACING = 16       # pt  (leading)
DEFAULT_MARGIN_MM = 20          # mm
DEFAULT_TEXT_COLOR = "#1a1a1a"
DEFAULT_BG_COLOR = "#ffffff"
DEFAULT_PAGE_SIZE = "A4"

# Built-in invitation template.  Override with --template <file>.
DEFAULT_TEMPLATE = """\
Hallo {name},

60 & 60 – ein guter Grund, das Leben zu feiern!

Das möchten wir gerne gemeinsam mit euch und laden herzlich am 22.05.2026 \
ab 18 Uhr in den Oberkellner in Greven ein.

Über eine Zusage freuen wir uns persönlich, telefonisch oder per Mail – \
am einfachsten aber über den Code {invite_code} unter www.zerfowski.de.
Dort könnt ihr auch Musikwünsche hinterlassen und die Party mitgestalten.

Mit 60 haben wir schon alles, daher seid ihr Geschenkebefreit.
Wer dennoch eine kleine Freude machen möchte, darf gerne das kleine \
Sparschwein füttern.

Wir freuen uns über jede positive Rückmeldung und vor allem auf einen \
fröhlichen Abend mit euch.

Liebe Grüße
    Birgit und Olaf

Olaf-und-Birgit@zerfowski.de
Mobil Olaf:   01578 123456
Mobil Birgit: 0176 123456

Anmeldecode www.zerfowski.de: {invite_code}
Anmeldelink:  {invite_link}
"""

# Regex to identify an invite-hash column (SHA-1 / SHA-256 hex string)
_HASH_RE = re.compile(r"^[0-9a-f]{40,}$")


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def parse_hex_color(hex_str: str) -> "colors.Color":
    """Convert '#RRGGBB' to a reportlab Color object."""
    h = hex_str.lstrip("#")
    if len(h) != 6:
        raise ValueError(
            f"Ungültige Farbe: {hex_str!r}  –  erwartet Format #RRGGBB."
        )
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return colors.Color(r / 255.0, g / 255.0, b / 255.0)


def get_page_size(name: str):
    """Return a reportlab page-size tuple for 'A4' or 'Letter'."""
    return {"a4": A4, "letter": letter}.get(name.lower(), A4)


def safe_filename(text: str) -> str:
    """Replace characters that are unsafe in filenames with underscores."""
    return re.sub(r"[^\w]", "_", text, flags=re.UNICODE).strip("_")


def normalize_header(raw) -> str:
    """Lower-case and strip the 'max_persons: 105' annotation if present."""
    if raw is None:
        return ""
    return str(raw).split(":")[0].strip().lower()


# ---------------------------------------------------------------------------
# Excel reading
# ---------------------------------------------------------------------------

# Maps a normalised header name to an internal field key.
_HEADER_MAP = {
    "name": "name",
    "nachname": "nachname",
    "kategorie": "kategorie",
    "max_persons": "max_persons",
    "invite_code": "invite_code",
    "email": "email",
    "telephone": "telephone",
    "notify_admin": "notify_admin",
    "text": "text",
}


def _build_col_map(header_row: tuple) -> dict[int, str]:
    """
    Build a mapping {column_index: internal_key} from the header row.

    Rules
    -----
    * Named columns are matched case-insensitively (see _HEADER_MAP).
    * If 'name' appears a second time it is mapped to 'full_name'
      (the second occurrence usually contains the fully formatted name).
    * The hash column has no header label – it is auto-detected later
      from the first data row.
    """
    col_map: dict[int, str] = {}
    seen_name = False

    for idx, cell in enumerate(header_row):
        norm = normalize_header(cell)
        if norm in _HEADER_MAP:
            key = _HEADER_MAP[norm]
            if key == "name" and seen_name:
                col_map[idx] = "full_name"
            else:
                col_map[idx] = key
                if key == "name":
                    seen_name = True

    return col_map


def _detect_hash_col(col_map: dict[int, str], first_data_row: tuple) -> dict[int, str]:
    """
    Scan the first data row for a long hex string in an unmapped column and
    register it as 'invite_hash'.
    """
    for idx, val in enumerate(first_data_row):
        if idx not in col_map and val and _HASH_RE.match(str(val).strip()):
            col_map[idx] = "invite_hash"
            break
    return col_map


def read_excel(excel_path: str, sheet_name: str) -> list[dict]:
    """
    Read *sheet_name* from *excel_path* and return a list of row dicts.
    Empty rows are skipped.
    """
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    except FileNotFoundError:
        sys.exit(f"Excel-Datei nicht gefunden: {excel_path}")
    except Exception as exc:
        sys.exit(f"Fehler beim Öffnen der Excel-Datei: {exc}")

    if sheet_name not in wb.sheetnames:
        sys.exit(
            f"Tabellenblatt '{sheet_name}' nicht gefunden.\n"
            f"Verfügbare Blätter: {', '.join(wb.sheetnames)}"
        )

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        sys.exit("Das Tabellenblatt ist leer.")

    header_row = rows[0]
    col_map = _build_col_map(header_row)

    if len(rows) > 1:
        col_map = _detect_hash_col(col_map, rows[1])

    records: list[dict] = []
    for raw_row in rows[1:]:
        if all(v is None for v in raw_row):
            continue  # skip blank rows
        record: dict = {}
        for idx, key in col_map.items():
            val = raw_row[idx] if idx < len(raw_row) else None
            record[key] = str(val).strip() if val is not None else ""
        records.append(record)

    return records


def show_column_mapping(excel_path: str, sheet_name: str) -> None:
    """Print the detected column mapping and exit (--show-columns)."""
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_row = rows[0] if rows else ()
    col_map = _build_col_map(header_row)
    if len(rows) > 1:
        col_map = _detect_hash_col(col_map, rows[1])

    print(f"\nSpalten-Mapping für: {excel_path}  |  Blatt: {sheet_name}\n")
    print(f"  {'Index':>5}  {'Header-Wert':<30}  {'Internes Feld'}")
    print("  " + "-" * 60)
    for idx, cell in enumerate(header_row):
        internal = col_map.get(idx, "– (ignoriert)")
        print(f"  {idx:>5}  {str(cell or ''):<30}  {internal}")
    print()


# ---------------------------------------------------------------------------
# Filtering
# ---------------------------------------------------------------------------

def apply_filters(records: list[dict], args: argparse.Namespace) -> list[dict]:
    """Return the subset of records that pass all active filters."""
    result = records

    if args.filter_kategorie:
        cats = {c.strip() for c in args.filter_kategorie}
        result = [r for r in result if r.get("kategorie", "") in cats]

    if args.filter_has_email:
        result = [r for r in result if r.get("email", "")]

    if args.filter_no_email:
        result = [r for r in result if not r.get("email", "")]

    return result


# ---------------------------------------------------------------------------
# PDF generation
# ---------------------------------------------------------------------------

def _build_invite_link(base_url: str, event_id: int, invite_hash: str) -> str:
    return f"{base_url.rstrip('/')}/event/{event_id}/invite/{invite_hash}"


def _draw_cover_page(
    c: "rl_canvas.Canvas",
    image_path: str,
    page_w: float,
    page_h: float,
    fit_mode: str,
) -> None:
    """
    Draw *image_path* on a blank page.

    fit_mode
    --------
    'fill'  – scale to cover the full page (may crop edges)
    'fit'   – scale to fit entirely within the page (may leave white space)
    """
    img = ImageReader(image_path)
    iw, ih = img.getSize()

    if fit_mode == "fill":
        scale = max(page_w / iw, page_h / ih)
    else:  # 'fit'
        scale = min(page_w / iw, page_h / ih)

    draw_w = iw * scale
    draw_h = ih * scale
    x = (page_w - draw_w) / 2
    y = (page_h - draw_h) / 2

    c.drawImage(image_path, x, y, width=draw_w, height=draw_h)


def _wrap_line(text: str, font_name: str, font_size: float, max_width: float, c) -> list[str]:
    """
    Word-wrap a single line of *text* to fit within *max_width* points.
    Returns a list of output lines.
    """
    if not text.strip():
        return [""]

    words = text.split(" ")
    lines: list[str] = []
    current = ""

    for word in words:
        candidate = (current + " " + word).strip()
        if c.stringWidth(candidate, font_name, font_size) <= max_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = word

    if current:
        lines.append(current)

    return lines if lines else [""]


def _draw_text_page(
    c: "rl_canvas.Canvas",
    text: str,
    page_w: float,
    page_h: float,
    cfg: dict,
) -> None:
    """
    Draw the invitation text on a single page (or overflow to additional pages
    if the text is longer than one page).
    """
    bg_color: "colors.Color" = cfg["bg_color"]
    text_color: "colors.Color" = cfg["text_color"]
    font_name: str = cfg["font_name"]
    font_size: float = cfg["font_size"]
    leading: float = cfg["line_spacing"]
    margin: float = cfg["margin"]

    text_w = page_w - 2 * margin
    text_h = page_h - 2 * margin

    def _new_page() -> float:
        """Start a new page and return the initial y position."""
        c.showPage()
        _fill_bg(c, bg_color, page_w, page_h)
        c.setFillColor(text_color)
        c.setFont(font_name, font_size)
        return page_h - margin - font_size

    def _fill_bg(canvas_obj, bg, pw, ph):
        canvas_obj.setFillColor(bg)
        canvas_obj.rect(0, 0, pw, ph, fill=1, stroke=0)

    # Draw background on the current (second) page
    _fill_bg(c, bg_color, page_w, page_h)
    c.setFillColor(text_color)
    c.setFont(font_name, font_size)

    y = page_h - margin - font_size

    for paragraph in text.split("\n"):
        wrapped = _wrap_line(paragraph, font_name, font_size, text_w, c)
        for output_line in wrapped:
            if y < margin:
                y = _new_page()
            c.drawString(margin, y, output_line)
            y -= leading


def generate_pdf(
    record: dict,
    template: str,
    image_path: str | None,
    output_path: Path,
    page_size: tuple,
    cfg: dict,
) -> None:
    """Generate a two-page invitation PDF for one guest record."""
    page_w, page_h = page_size

    # Build placeholder dict from the record
    invite_hash = record.get("invite_hash", "")
    invite_link = (
        _build_invite_link(cfg["base_url"], cfg["event_id"], invite_hash)
        if invite_hash
        else ""
    )

    full_name = (
        record.get("full_name")
        or f"{record.get('name', '')} {record.get('nachname', '')}".strip()
    )

    placeholders = {
        "name": record.get("name", ""),
        "nachname": record.get("nachname", ""),
        "full_name": full_name,
        "invite_code": record.get("invite_code", ""),
        "max_persons": record.get("max_persons", ""),
        "invite_link": invite_link,
        "email": record.get("email", ""),
        "telephone": record.get("telephone", ""),
    }

    # Resolve the invitation text
    if cfg.get("use_text_column") and record.get("text"):
        invite_text = record["text"]
    else:
        try:
            invite_text = template.format_map(placeholders)
        except KeyError as exc:
            print(
                f"  Warnung: Unbekannter Platzhalter {exc} im Template – "
                "wird leer gelassen.",
                file=sys.stderr,
            )
            # Fall back to safe substitution
            class _SafeDict(dict):
                def __missing__(self, key):
                    return f"{{{key}}}"

            invite_text = template.format_map(_SafeDict(placeholders))

    # Build the PDF
    c = rl_canvas.Canvas(str(output_path), pagesize=page_size)

    # --- Page 1: cover image (or blank) ---
    if image_path:
        _draw_cover_page(c, image_path, page_w, page_h, cfg.get("image_fit", "fill"))
    else:
        c.setFillColor(colors.white)
        c.rect(0, 0, page_w, page_h, fill=1, stroke=0)

    c.showPage()

    # --- Page 2: invitation text ---
    _draw_text_page(c, invite_text, page_w, page_h, cfg)
    c.showPage()

    c.save()


# ---------------------------------------------------------------------------
# Argument parser
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="generate_invitations.py",
        description="Einladungs-PDFs aus einer Excel-Gästeliste erstellen.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )

    # ---- Input / Output ----
    io = p.add_argument_group("Eingabe / Ausgabe")
    io.add_argument(
        "--excel",
        default=DEFAULT_EXCEL,
        metavar="DATEI",
        help=f"Excel-Datei  (Standard: {DEFAULT_EXCEL})",
    )
    io.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
        metavar="BLATT",
        help=f"Tabellenblatt  (Standard: {DEFAULT_SHEET})",
    )
    io.add_argument(
        "--image",
        metavar="BILD.jpg",
        help="JPG/PNG-Bild für Seite 1 der Einladung",
    )
    io.add_argument(
        "--template",
        metavar="DATEI",
        help="Textdatei mit Einladungsvorlage und {Platzhaltern}. "
             "Wird kein Template angegeben, wird das eingebaute Standard-Template verwendet.",
    )
    io.add_argument(
        "--output-dir",
        default=DEFAULT_OUTPUT_DIR,
        metavar="VERZEICHNIS",
        help=f"Ausgabe-Verzeichnis für die PDFs  (Standard: {DEFAULT_OUTPUT_DIR})",
    )
    io.add_argument(
        "--use-text-column",
        action="store_true",
        help="Text direkt aus der Excel-Spalte 'Text' verwenden "
             "(statt Template + Platzhalter)",
    )

    # ---- Design ----
    design = p.add_argument_group("Design (Seite 2 – Einladungstext)")
    design.add_argument(
        "--font",
        metavar="DATEI.ttf",
        help="Pfad zu einer TTF-Schriftdatei.  "
             "Ohne diese Option wird Helvetica verwendet.",
    )
    design.add_argument(
        "--font-size",
        type=float,
        default=DEFAULT_FONT_SIZE,
        metavar="PT",
        help=f"Schriftgröße in Punkt  (Standard: {DEFAULT_FONT_SIZE})",
    )
    design.add_argument(
        "--line-spacing",
        type=float,
        default=DEFAULT_LINE_SPACING,
        metavar="PT",
        help=f"Zeilenabstand in Punkt  (Standard: {DEFAULT_LINE_SPACING})",
    )
    design.add_argument(
        "--text-color",
        default=DEFAULT_TEXT_COLOR,
        metavar="#RRGGBB",
        help=f"Schriftfarbe als Hex-Wert  (Standard: {DEFAULT_TEXT_COLOR})",
    )
    design.add_argument(
        "--bg-color",
        default=DEFAULT_BG_COLOR,
        metavar="#RRGGBB",
        help=f"Hintergrundfarbe als Hex-Wert  (Standard: {DEFAULT_BG_COLOR})",
    )
    design.add_argument(
        "--margin",
        type=float,
        default=DEFAULT_MARGIN_MM,
        metavar="MM",
        help=f"Seitenrand in Millimetern  (Standard: {DEFAULT_MARGIN_MM})",
    )
    design.add_argument(
        "--page-size",
        default=DEFAULT_PAGE_SIZE,
        choices=["A4", "Letter"],
        metavar="FORMAT",
        help=f"Seitenformat: A4 oder Letter  (Standard: {DEFAULT_PAGE_SIZE})",
    )
    design.add_argument(
        "--image-fit",
        default="fill",
        choices=["fill", "fit"],
        help="Bildanpassung Seite 1: "
             "'fill' = Seite vollständig füllen (kann Bildränder abschneiden), "
             "'fit'  = Bild vollständig anzeigen mit möglichen Rändern  (Standard: fill)",
    )

    # ---- Invite link ----
    link = p.add_argument_group("Einladungslink-Konstruktion")
    link.add_argument(
        "--base-url",
        default=DEFAULT_BASE_URL,
        metavar="URL",
        help=f"Basis-URL für Einladungslinks  (Standard: {DEFAULT_BASE_URL})",
    )
    link.add_argument(
        "--event-id",
        type=int,
        default=DEFAULT_EVENT_ID,
        metavar="N",
        help=f"Event-ID für Einladungslinks  (Standard: {DEFAULT_EVENT_ID})",
    )

    # ---- Filter ----
    flt = p.add_argument_group("Filter")
    flt.add_argument(
        "--filter-kategorie",
        nargs="+",
        metavar="KAT",
        help="Nur Gäste der angegebenen Kategorie(n), z. B. --filter-kategorie Familie Freunde",
    )
    flt.add_argument(
        "--filter-has-email",
        action="store_true",
        help="Nur Gäste MIT E-Mail-Adresse verarbeiten",
    )
    flt.add_argument(
        "--filter-no-email",
        action="store_true",
        help="Nur Gäste OHNE E-Mail-Adresse verarbeiten",
    )

    # ---- Misc ----
    misc = p.add_argument_group("Sonstiges")
    misc.add_argument(
        "--show-columns",
        action="store_true",
        help="Erkannte Spalten-Zuordnung der Excel-Tabelle anzeigen und beenden",
    )
    misc.add_argument(
        "--dry-run",
        action="store_true",
        help="Zeigt welche PDFs erzeugt würden, ohne Dateien zu erstellen",
    )
    misc.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Ausführliche Ausgabe",
    )

    return p


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    # ---- Resolve Excel path ----
    excel_path = Path(args.excel)
    if not excel_path.exists():
        sys.exit(f"Excel-Datei nicht gefunden: {excel_path}")

    # ---- --show-columns: just print column mapping and exit ----
    if args.show_columns:
        show_column_mapping(str(excel_path), args.sheet)
        return

    # ---- Image ----
    image_path: str | None = None
    if args.image:
        img = Path(args.image)
        if not img.exists():
            sys.exit(f"Bild-Datei nicht gefunden: {args.image}")
        image_path = str(img.resolve())
    else:
        print(
            "Hinweis: Kein --image angegeben – Seite 1 bleibt leer.",
            file=sys.stderr,
        )

    # ---- Template ----
    if args.use_text_column:
        template = ""  # will be ignored; text comes from 'text' column
    elif args.template:
        tpl = Path(args.template)
        if not tpl.exists():
            sys.exit(f"Template-Datei nicht gefunden: {args.template}")
        template = tpl.read_text(encoding="utf-8")
        if args.verbose:
            print(f"Template geladen: {tpl}")
    else:
        template = DEFAULT_TEMPLATE
        if args.verbose:
            print("Verwende eingebautes Standard-Template.")

    # ---- Font ----
    font_name = "Helvetica"
    if args.font:
        font_path = Path(args.font)
        if not font_path.exists():
            sys.exit(f"Schriftdatei nicht gefunden: {args.font}")
        font_name = "InvitationFont"
        try:
            pdfmetrics.registerFont(TTFont(font_name, str(font_path)))
            if args.verbose:
                print(f"Schrift registriert: {font_path.name}")
        except Exception as exc:
            sys.exit(f"Schriftdatei konnte nicht geladen werden: {exc}")

    # ---- Colors ----
    try:
        text_color = parse_hex_color(args.text_color)
        bg_color = parse_hex_color(args.bg_color)
    except ValueError as exc:
        sys.exit(str(exc))

    # ---- Page setup ----
    page_size = get_page_size(args.page_size)
    margin_pt = args.margin * mm

    cfg = {
        "font_name": font_name,
        "font_size": args.font_size,
        "line_spacing": args.line_spacing,
        "text_color": text_color,
        "bg_color": bg_color,
        "margin": margin_pt,
        "base_url": args.base_url,
        "event_id": args.event_id,
        "image_fit": args.image_fit,
        "use_text_column": args.use_text_column,
    }

    # ---- Read Excel ----
    print(f"Lese Excel: {excel_path}  (Blatt: '{args.sheet}')")
    records = read_excel(str(excel_path), args.sheet)
    print(f"  {len(records)} Zeile(n) gefunden.")

    # ---- Apply filters ----
    records = apply_filters(records, args)
    if not records:
        print("Keine Einträge nach Filterung – fertig.")
        return
    print(f"  {len(records)} Einträge nach Filterung.")

    # ---- Output directory ----
    output_dir = Path(args.output_dir)
    if not args.dry_run:
        output_dir.mkdir(parents=True, exist_ok=True)

    # ---- Generate PDFs ----
    generated = 0
    skipped = 0

    for record in records:
        name = record.get("name", "unbekannt")
        invite_code = record.get("invite_code", "")

        if not invite_code:
            print(
                f"  Überspringe '{name}': kein invite_code vorhanden.",
                file=sys.stderr,
            )
            skipped += 1
            continue

        filename = f"{safe_filename(name)}_{invite_code}.pdf"
        output_path = output_dir / filename

        if args.dry_run:
            full_name = record.get("full_name") or f"{name} {record.get('nachname', '')}".strip()
            print(f"  [dry-run] {output_path}  ({full_name})")
            generated += 1
            continue

        if args.verbose:
            print(f"  Erstelle: {output_path}")

        try:
            generate_pdf(record, template, image_path, output_path, page_size, cfg)
            generated += 1
        except Exception as exc:
            print(
                f"  Fehler bei '{name}' ({invite_code}): {exc}",
                file=sys.stderr,
            )
            skipped += 1

    # ---- Summary ----
    action = "würden erstellt werden" if args.dry_run else "erstellt"
    print(
        f"\nFertig: {generated} PDF(s) {action}, {skipped} übersprungen."
    )
    if not args.dry_run and generated:
        print(f"Ausgabe-Verzeichnis: {output_dir.resolve()}")


if __name__ == "__main__":
    main()
